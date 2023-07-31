import openpyxl
import argparse
from openpyxl.styles import Alignment, PatternFill, Font

colors = ["c1bcbc", "ffd92f", "76cf73", "519fd5", "ef6e69", "9943bd"]

header = ["Patients", "Treatments before V1", "Ongoing treatments at V1", "Ongoing treatments at V2",
          "Treatments after last visit", "Treatments after V1 and before V2"]
header = [x.upper() for x in header]

# Define the fill pattern to use for highlighting cells containing "Missing"
missing_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

def create_header(sheet, i):
    header_coordinates = [cell.column_letter+"1" for cell in headers0[i]]
    sheet.merge_cells(':'.join(header_coordinates))
    sheet[header_coordinates[0]] = header[i]

# Parse the filename argument from the command line
parser = argparse.ArgumentParser()
parser.add_argument("filename", help="the name of the Excel file to merge empty cells in")
args = parser.parse_args()

# Load the Excel file using openpyxl
wb = openpyxl.load_workbook(args.filename)

# Iterate over all sheets in the workbook
for sheet in wb:
    # Find column indices by header name
    headers = sheet[1]
    col_date_1st_visit = None
    col_date_2nd_visit = None
    col_followup = None
    col_last_beginning = None

    for cell in headers:
        if "Date of the 1st visit" == str(cell.value):
            col_date_1st_visit = cell.column-1
        elif "Date of the 2nd visit" == str(cell.value):
            col_date_2nd_visit = cell.column-1
        elif "Date of the follow-up" == str(cell.value):
            col_followup  = cell.column-1
        elif "Date of beginning" == str(cell.value):
            col_last_beginning  = cell.column-1

    headers0 = [[headers[0], headers[1]],
               [headers[2], headers[col_date_1st_visit-1]],
               [headers[col_date_1st_visit], headers[col_date_2nd_visit-1]],
               [headers[col_date_2nd_visit], headers[col_followup-1]],
               [headers[col_followup], headers[col_last_beginning-1]],
               [headers[col_last_beginning], headers[-1]]]
    header_letters = [[cell.column_letter for cell in row] for row in headers0]

    # Define the colors to use for each column
    column_colors = dict(zip([':'.join(row) for row in header_letters], colors))

    # Add a first row with the specified merged cells
    sheet.insert_rows(1)
    for i in range(len(headers0)):
        create_header(sheet, i)

    # Color each column with the corresponding color
    for range_str, color in column_colors.items():
        for cell in sheet[range_str]:
            if isinstance(cell, tuple):
                cell_list = cell
            else:
                cell_list = [cell]
            for c in cell_list:
                c.fill = PatternFill(fgColor=color, fill_type="solid")

    # Iterate over all columns in the sheet
    for col in sheet.columns:
        # Highlight cells containing the word "Missing" in yellow
        for cell in col:
            if cell.value == "Missing":
                cell.fill = missing_fill
            # Replace any cell containing "NA" with an empty string
            if cell.value == "NA":
                cell.value = ""

    # Make the header row bold
    for cell in sheet[1]:
        cell.font = Font(name='Calibri', bold=True)
    for cell in sheet[2]:
        cell.font = Font(name='Calibri', bold=True, italic=True)

# Save the modified Excel file with the same name as the original file
wb.save(args.filename)
